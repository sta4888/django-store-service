import json
import logging

import requests
from django.core.cache import cache
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from core.settings import FASTAPI_SERVICE_URL
# from .tasks import update_user_stats_cache
from django.http import JsonResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from .models import Purchase, News, MonthlyReport
from datetime import datetime, timedelta
from accounts.models import CustomUser

from .services.fastapi_service import FastAPIService

logger = logging.getLogger(__name__)




@login_required
def dashboard_view(request):
    stats = {}
    # user_id = request.user.username
    # cache_key = f"user:stats:{user_id}"
    # stats = cache.get(cache_key)  # быстрое чтение, не блокирующее
    #
    # if stats is None:
    #     # стартуем фоновую задачу, но не ждём её результата
    #     update_user_stats_cache.delay(user_id)

    return render(
        request,
        "cabinet/dashboard.html",
        {
            "user_stats": stats or {},  # пустой словарь пока нет данных
            "loading": stats is None,
        }
    )


@login_required
def profile_view(request):
    return render(request, 'cabinet/profile.html')


@login_required
def purchases_view(request):
    purchases = Purchase.objects.filter(user=request.user).order_by('-date')
    return render(request, 'cabinet/purchases.html', {'purchases': purchases})


@login_required
def finance_view(request):
    return render(request, 'cabinet/finance.html')



@login_required
def get_user_data(request):
    user_id = request.user.username
    try:
        r = requests.get(
            f"{FASTAPI_SERVICE_URL}/user/users/{user_id}/status",
            timeout=5
        )
        r.raise_for_status()
    except requests.RequestException as e:
        return JsonResponse(
            {"error": str(e)},
            status=503
        )

    return JsonResponse(r.json(), safe=False)



@login_required
def add_user_lo(request):
    # user_id = request.user.username

    # Получаем данные из POST запроса
    if request.method == 'POST':
        try:
            # Пытаемся получить данные из JSON
            data = json.loads(request.body)
        except json.JSONDecodeError:
            # Если не JSON, пробуем получить из формы
            data = request.POST.dict()

        lo_amount = data.get('lo')
        user_id = data.get('user_id')

        if not lo_amount:
            return JsonResponse(
                {"error": "Параметр 'lo' не указан"},
                status=400
            )

        try:
            payload = {"lo": float(lo_amount)}

            r = requests.post(
                f"{FASTAPI_SERVICE_URL}/user/users/{user_id}/lo/add",
                json=payload,
                timeout=5
            )
            r.raise_for_status()

            return JsonResponse(r.json(), safe=False)

        except ValueError:
            return JsonResponse(
                {"error": "Параметр 'lo' должен быть числом"},
                status=400
            )
        except requests.RequestException as e:
            return JsonResponse(
                {"error": str(e)},
                status=503
            )

    # Если метод не POST
    return JsonResponse(
        {"error": "Метод не поддерживается. Используйте POST."},
        status=405
    )


@login_required
def sub_user_lo(request):
    # user_id = request.user.username

    # Получаем данные из POST запроса
    if request.method == 'POST':
        try:
            # Пытаемся получить данные из JSON
            data = json.loads(request.body)
        except json.JSONDecodeError:
            # Если не JSON, пробуем получить из формы
            data = request.POST.dict()

        lo_amount = data.get('lo')
        user_id = data.get('user_id')

        if not lo_amount:
            return JsonResponse(
                {"error": "Параметр 'lo' не указан"},
                status=400
            )

        try:
            payload = {"lo": float(lo_amount)}

            r = requests.post(
                f"{FASTAPI_SERVICE_URL}/user/users/{user_id}/lo/subtract",
                json=payload,
                timeout=5
            )
            r.raise_for_status()

            return JsonResponse(r.json(), safe=False)

        except ValueError:
            return JsonResponse(
                {"error": "Параметр 'lo' должен быть числом"},
                status=400
            )
        except requests.RequestException as e:
            return JsonResponse(
                {"error": str(e)},
                status=503
            )

    # Если метод не POST
    return JsonResponse(
        {"error": "Метод не поддерживается. Используйте POST."},
        status=405
    )


@login_required
def get_user_team(request):
    user_id = request.user.username
    try:
        r = requests.get(
            f"{FASTAPI_SERVICE_URL}/user/users/{user_id}/structure",
            timeout=5
        )
        r.raise_for_status()
    except requests.RequestException as e:
        return JsonResponse(
            {"error": str(e)},
            status=503
        )

    return JsonResponse(r.json(), safe=False)


@login_required
def api_test_page(request):
    """Страница для тестирования API"""
    return render(request, 'cabinet/admin_api.html')


@login_required
def admin_panel(request):
    """Админ-панель для магазинов и администраторов"""
    if not request.user.can_access_admin:
        raise PermissionDenied("У вас нет доступа к админ-панели")

    return render(request, 'cabinet/admin_panel.html')


@login_required
def structure_view(request):
    """Страница структуры рефералов пользователя"""
    from datetime import date

    user = request.user

    # Получаем рефералов первого уровня
    direct_referrals = CustomUser.objects.filter(
        referrer=user
    ).select_related('referrer').order_by('-date_joined')

    # Пагинация
    paginator = Paginator(direct_referrals, 20)  # 20 записей на страницу
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Получаем общую статистику
    total_referrals = user.total_referrals
    active_referrals = user.active_referrals
    group_volume = user.group_volume

    # Проверяем, был ли уже сохранён отчёт за текущий месяц
    today = date.today()
    report_already_saved = MonthlyReport.objects.filter(
        user=user,
        year=today.year,
        month=today.month,
    ).exists()

    # Дата следующего доступного сохранения — 1-е число следующего месяца
    if today.month == 12:
        next_available_date = date(today.year + 1, 1, 1)
    else:
        next_available_date = date(today.year, today.month + 1, 1)

    context = {
        'direct_referrals': page_obj,
        'total_referrals': total_referrals,
        'active_referrals': active_referrals,
        'group_volume': group_volume,
        'report_already_saved': report_already_saved,
        'next_available_date': next_available_date,
    }

    return render(request, 'cabinet/structure.html', context)


# Дополнительная функция для AJAX-загрузки рефералов (по желанию)
@login_required
def get_referrals_json(request):
    """Получение списка рефералов в формате JSON с данными из API"""
    user = request.user
    level = request.GET.get('level', '1')
    
    try:
        level = int(level)
    except ValueError:
        level = 1
    
    try:
        # Получаем структуру из FastAPI
        response = requests.get(
            f"{FASTAPI_SERVICE_URL}/user/users/{user.username}/structure",
            timeout=5
        )
        response.raise_for_status()
        api_data = response.json()
        print(f"api_data.get('error') -- {api_data.get('error')}")
        
        if api_data.get('error'):
            return JsonResponse({
                'error': True,
                'message': api_data.get('error_msg', 'Ошибка при получении данных')
            }, status=500)
            
        structure_data = api_data.get('data', {})
        team_members = structure_data.get('team', [])
        
    except requests.RequestException as e:
        logger.error(f"Ошибка API при получении структуры: {e}")
        return JsonResponse({
            'error': True,
            'message': str(e)
        }, status=503)
    
    # Получаем данные из БД
    referral_user_ids = [str(member.get('user_id')) for member in team_members]
    db_referrals = CustomUser.objects.filter(user_id__in=referral_user_ids)
    db_referrals_dict = {str(ref.user_id): ref for ref in db_referrals}
    
    # Обрабатываем только первый уровень
    if level == 1:
        referrals_data = []
        for api_member in team_members:
            user_id = str(api_member.get('user_id'))
            db_data = db_referrals_dict.get(user_id)
            
            if db_data:
                member_data = {
                    'id': user_id,
                    'name': db_data.get_full_name() or db_data.username,
                    'email': db_data.email,
                    'phone': db_data.phone,
                    'country': db_data.country,
                    'registration_date': db_data.date_joined.strftime('%d.%m.%Y'),
                    'personal_volume': float(api_member.get('lo', 0)),  # Из API
                    'group_volume': float(db_data.group_volume),  # Из БД
                    'partner_level': db_data.partner_level,
                    'user_type': db_data.user_type,
                    'total_referrals': db_data.total_referrals,
                    'active_referrals': db_data.active_referrals,
                    'earnings': float(db_data.earnings),
                    'team_count': len(api_member.get('team', [])),
                }
                referrals_data.append(member_data)
        
        return JsonResponse({
            'level': level,
            'referrals': referrals_data,
            'total_count': len(referrals_data),
            'error': False
        })
    
    # Для второго уровня и глубже (рекурсивно)
    elif level > 1:
        all_referrals_data = []
        
        def get_deep_referrals(members, current_level, max_level):
            if current_level > max_level:
                return []
            
            referrals_data = []
            for member in members:
                user_id = str(member.get('user_id'))
                
                try:
                    # Получаем структуру для каждого члена команды
                    member_response = requests.get(
                        f"{FASTAPI_SERVICE_URL}/user/users/{user_id}/structure",
                        timeout=3
                    )
                    if member_response.status_code == 200:
                        member_data = member_response.json()
                        if not member_data.get('error'):
                            # Получаем данные из БД
                            try:
                                db_data = CustomUser.objects.get(user_id=user_id)
                            except CustomUser.DoesNotExist:
                                db_data = None
                            
                            member_info = {
                                'id': user_id,
                                'level': current_level,
                                'personal_volume': float(member.get('lo', 0)),
                                'team_count': len(member.get('team', [])),
                            }
                            
                            if db_data:
                                member_info.update({
                                    'name': db_data.get_full_name() or db_data.username,
                                    'email': db_data.email,
                                    'phone': db_data.phone,
                                    'registration_date': db_data.date_joined.strftime('%d.%m.%Y'),
                                    'group_volume': float(db_data.group_volume),
                                    'partner_level': db_data.partner_level,
                                })
                            
                            referrals_data.append(member_info)
                            
                            # Рекурсивно получаем команду
                            if current_level < max_level:
                                sub_team = member_data.get('data', {}).get('team', [])
                                referrals_data.extend(
                                    get_deep_referrals(sub_team, current_level + 1, max_level)
                                )
                except:
                    continue
            
            return referrals_data
        
        # Начинаем рекурсию с команды первого уровня
        deep_referrals = get_deep_referrals(team_members, 2, level)
        
        return JsonResponse({
            'level': level,
            'referrals': deep_referrals,
            'total_count': len(deep_referrals),
            'error': False
        })
    
    return JsonResponse({
        'level': level,
        'referrals': [],
        'total_count': 0,
        'error': False
    })



from datetime import date

EXTRA_BONUS_RANK = {
    "-":                     0,
    "Par dazmol":            1,
    "Changyutgich":          2,
    "Televizor smart-32":    3,
    "Gaz plita":             4,
    "Konditsioner":          5,
    "Kir yuvadigan mashina": 6,
    "Chet el sayohati":      7,
    "Onix avtomobili":       8,
    "Chery Tigo 7 Pro Max":  9,
}


def get_prev_month(today):
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


@login_required
def referral_tree_api(request):
    """
    Возвращает дерево рефералов.
    Для каждого узла параллельно запрашивает /status из FastAPI
    чтобы получить актуальные lo, go, qualification и бонусы.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    nodes = []
    edges = []
    visited = set()

    # ── Прошлый месяц ──────────────────────────────────────────────
    today = date.today()
    prev_year, prev_month = get_prev_month(today)

    # Берём extra_bonus каждого пользователя за прошлый месяц
    prev_extra_bonus_map = dict(
        MonthlyReport.objects.filter(
            year=prev_year,
            month=prev_month,
        ).exclude(
            extra_bonus=''
        ).values_list('user_id', 'extra_bonus')
    )

    # ── 1. Обходим дерево Django БД, собираем базовые данные ───────
    def get_short_name(user):
        parts = [user.first_name, user.last_name]
        name = ' '.join(p for p in parts if p).strip()
        return name or user.username

    def get_full_name(user):
        parts = [user.last_name, user.first_name, getattr(user, 'middle_name', '')]
        name = ' '.join(p for p in parts if p).strip()
        return name or user.username

    def traverse(user, level):
        if user.pk in visited:
            return
        visited.add(user.pk)

        nodes.append({
            'id':              user.user_id,
            'label':           get_short_name(user),
            'title':           get_full_name(user),
            'level':           level,
            'active':          getattr(user, 'is_active', True),
            'user_type':       getattr(user, 'user_type', 'partner'),
            'personal_volume': 0.0,
            'group_volume':    0.0,
            'partner_level':   getattr(user, 'partner_level', ''),
            'total_referrals': getattr(user, 'total_referrals', 0),
            'qualification':   '',
            'side_volume':     0,
            'points':          0,
            'personal_bonus':  0,
            'structure_bonus': 0,
            'mentor_bonus':    0,
            'extra_bonus':     '',
            'personal_money':  0,
            'group_money':     0,
            'leader_money':    0,
            'side_vol_money':  0,
            'total_money':     0,
            'veron':           0,
            'total_income':    0,
        })

        referrals = CustomUser.objects.filter(referrer=user)
        for referral in referrals:
            edges.append({'from': user.user_id, 'to': referral.user_id})
            traverse(referral, level + 1)

    traverse(request.user, level=0)

    # ── 2. Параллельно запрашиваем /status для каждого узла ────────
    def fetch_status(user_id):
        try:
            resp = requests.get(
                f"{FASTAPI_SERVICE_URL}/user/users/{user_id}/status",
                timeout=5
            )
            if resp.status_code == 200:
                data = resp.json()
                if not data.get('error'):
                    return user_id, data.get('data', {})
        except Exception as e:
            logger.warning(f"Не удалось получить статус для {user_id}: {e}")
        return user_id, {}

    user_ids = [n['id'] for n in nodes]
    status_map = {}

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_status, uid): uid for uid in user_ids}
        for future in as_completed(futures):
            uid, status = future.result()
            status_map[uid] = status

    # ── 3. Обогащаем узлы данными из FastAPI ───────────────────────
    for node in nodes:
    st = status_map.get(node['id'], {})

    current_extra = st.get('extra_bonus', '') or ''
    prev_extra    = prev_extra_bonus_map.get(node['id'], '')
    current_rank  = EXTRA_BONUS_RANK.get(current_extra, 0)
    prev_rank     = EXTRA_BONUS_RANK.get(prev_extra, 0)

    # ── ОТЛАДКА — удалить после проверки ──
    logger.warning(
        f"USER {node['id']} | "
        f"current_extra='{current_extra}' (rank={current_rank}) | "
        f"prev_extra='{prev_extra}' (rank={prev_rank}) | "
        f"show={current_rank > prev_rank}"
    )
    # ──────────────────────────────────────

    node['extra_bonus'] = current_extra if current_rank > prev_rank else ''
    for node in nodes:
        
        st = status_map.get(node['id'], {})

        # Проверяем extra_bonus
        current_extra = st.get('extra_bonus', '') or ''
        prev_extra    = prev_extra_bonus_map.get(node['id'], '')
        current_rank  = EXTRA_BONUS_RANK.get(current_extra, 0)
        prev_rank     = EXTRA_BONUS_RANK.get(prev_extra, 0)
        # Показываем только если текущий бонус ВЫШЕ прошлого
        node['extra_bonus'] = current_extra if current_rank > prev_rank else ''

        if st:
            node['personal_volume'] = float(st.get('lo', 0) or 0)
            node['group_volume']    = float(st.get('go', 0) or 0)
            node['qualification']   = st.get('qualification', '')
            node['partner_level']   = st.get('qualification', node['partner_level'])
            node['side_volume']     = st.get('side_volume', 0)
            node['points']          = st.get('points', 0)
            node['personal_bonus']  = st.get('personal_bonus', 0)
            node['structure_bonus'] = st.get('structure_bonus', 0)
            node['mentor_bonus']    = st.get('mentor_bonus', 0)
            node['personal_money']  = st.get('personal_money', 0)
            node['group_money']     = st.get('group_money', 0)
            node['leader_money']    = st.get('leader_money', 0)
            node['side_vol_money']  = st.get('side_vol_money', 0)
            node['total_money']     = st.get('total_money', 0)
            node['veron']           = st.get('veron', 0)
            node['total_income']    = st.get('total_income', 0)

    return JsonResponse({'nodes': nodes, 'edges': edges})


@login_required
def generate_monthly_report(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

    from django.db.models import Sum, Count
    from concurrent.futures import ThreadPoolExecutor, as_completed

    today = date.today()
    year  = today.year
    month = today.month
    prev_year, prev_month = get_prev_month(today)

    # ── Прошлый месяц — для проверки extra_bonus ───────────────────
    prev_extra_bonus_map = dict(
        MonthlyReport.objects.filter(
            year=prev_year,
            month=prev_month,
        ).exclude(
            extra_bonus=''
        ).values_list('user_id', 'extra_bonus')
    )

    # ── 1. Обходим всё дерево структуры ────────────────────────────
    all_users = []
    visited   = set()

    def traverse(user):
        if user.pk in visited:
            return
        visited.add(user.pk)
        all_users.append(user)
        for referral in CustomUser.objects.filter(referrer=user):
            traverse(referral)

    traverse(request.user)

    # ── 2. Параллельно запрашиваем /status из FastAPI ──────────────
    def fetch_status(user):
        try:
            resp = requests.get(
                f"{FASTAPI_SERVICE_URL}/user/users/{user.username}/status",
                timeout=5
            )
            if resp.status_code == 200:
                data = resp.json()
                if not data.get('error'):
                    return user.pk, data.get('data', {})
        except Exception as e:
            logger.warning(f"Не удалось получить статус для {user.username}: {e}")
        return user.pk, {}

    status_map = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_status, u): u for u in all_users}
        for future in as_completed(futures):
            pk, status = future.result()
            status_map[pk] = status

    # ── 3. Сохраняем MonthlyReport ─────────────────────────────────
    saved  = 0
    errors = 0

    for user in all_users:
        st = status_map.get(user.pk, {})

        # Проверяем extra_bonus
        current_extra = st.get('extra_bonus', '') or ''
        prev_extra    = prev_extra_bonus_map.get(user.pk, '')
        current_rank  = EXTRA_BONUS_RANK.get(current_extra, 0)
        prev_rank     = EXTRA_BONUS_RANK.get(prev_extra, 0)
        # Сохраняем только если текущий бонус ВЫШЕ прошлого
        extra_bonus_value = current_extra if current_rank > prev_rank else ''

        purchases_agg = Purchase.objects.filter(
            user=user, date__year=year, date__month=month
        ).aggregate(
            count=Count('id'),
            amount=Sum('amount'),
        )

        new_referrals = CustomUser.objects.filter(
            referrer=user,
            date_joined__year=year,
            date_joined__month=month,
        ).count()

        try:
            MonthlyReport.objects.update_or_create(
                user=user,
                year=year,
                month=month,
                defaults={
                    'personal_volume':        st.get('lo', 0) or 0,
                    'group_volume':           st.get('go', 0) or 0,
                    'side_volume':            st.get('side_volume', 0) or 0,
                    'points':                 st.get('points', 0) or 0,
                    'veron':                  st.get('veron', 0) or 0,
                    'personal_bonus':         st.get('personal_bonus', 0) or 0,
                    'structure_bonus':        st.get('structure_bonus', 0) or 0,
                    'mentor_bonus':           st.get('mentor_bonus', 0) or 0,
                    'extra_bonus':            extra_bonus_value,
                    'bonus_total':            sum(float(st.get(k, 0) or 0) for k in ('personal_bonus', 'structure_bonus', 'mentor_bonus')),
                    'personal_money':         st.get('personal_money', 0) or 0,
                    'group_money':            st.get('group_money', 0) or 0,
                    'leader_money':           st.get('leader_money', 0) or 0,
                    'side_vol_money':         st.get('side_vol_money', 0) or 0,
                    'total_money':            st.get('total_money', 0) or 0,
                    'total_income':           st.get('total_income', 0) or 0,
                    'partner_level':          st.get('qualification', '') or '',
                    'new_referrals':          new_referrals,
                    'active_referrals_count': user.active_referrals,
                    'purchases_count':        purchases_agg['count'] or 0,
                    'purchases_amount':       purchases_agg['amount'] or 0,
                }
            )
            saved += 1
        except Exception as e:
            logger.error(f"Ошибка сохранения отчёта для {user.username}: {e}")
            errors += 1

    logger.info(f"MonthlyReport: сохранено {saved}, ошибок {errors} ({month}/{year})")

    # ── 4. Параллельно сбрасываем данные в FastAPI ─────────────────
    def reset_user_data(user):
        try:
            resp = requests.post(
                f"{FASTAPI_SERVICE_URL}/user/users/{user.username}/reset",
                timeout=5
            )
            if resp.status_code == 200:
                payload = resp.json()
                if not payload.get('error'):
                    return user.pk, True
        except Exception as e:
            logger.warning(f"Не удалось сбросить данные для {user.username}: {e}")
        return user.pk, False

    reset_ok     = 0
    reset_errors = 0

    with ThreadPoolExecutor(max_workers=10) as executor:
        reset_futures = {executor.submit(reset_user_data, u): u for u in all_users}
        for future in as_completed(reset_futures):
            _pk, success = future.result()
            if success:
                reset_ok += 1
            else:
                reset_errors += 1

    logger.info(f"FastAPI reset: успешно {reset_ok}, ошибок {reset_errors} ({month}/{year})")

    return JsonResponse({
        'success':      True,
        'saved':        saved,
        'errors':       errors,
        'month':        month,
        'year':         year,
        'reset_ok':     reset_ok,
        'reset_errors': reset_errors,
    })



def get_referral_details(request, user_id):
    """Получение детальной информации о реферале"""
    try:
        # Проверяем, что пользователь существует
        db_user = CustomUser.objects.get(user_id=user_id)
        
        # # Проверяем, что это реферал текущего пользователя
        # if db_user.referrer != request.user:
        #     return JsonResponse({'error': True, 'message': 'Доступ запрещен'}, status=403)
        
        # Получаем данные из API
        try:
            response = requests.get(
                f"{FASTAPI_SERVICE_URL}/user/users/{user_id}/status",
                timeout=5
            )
            if response.status_code == 200:
                api_data = response.json()
                lo_amount = api_data.get('lo', 0)
            else:
                lo_amount = 0
        except:
            lo_amount = 0
        
        data = {
            'id': db_user.user_id,
            'name': db_user.get_full_name() or db_user.username,
            'email': db_user.email,
            'phone': db_user.phone,
            'country': db_user.country,
            'registration_date': db_user.date_joined.strftime('%d.%m.%Y'),
            'personal_volume': float(lo_amount),  # Из API
            'group_volume': float(db_user.group_volume),
            'partner_level': db_user.partner_level,
            'user_type': db_user.user_type,
            'total_referrals': db_user.total_referrals,
            'active_referrals': db_user.active_referrals,
            'earnings': float(db_user.earnings),
            'available_for_withdrawal': float(db_user.available_for_withdrawal),
            'middle_name': db_user.middle_name or '',
            'passport_number': db_user.passport_number or '',
            'is_email_verified': db_user.is_email_verified,
            'is_terms_accepted': db_user.is_terms_accepted,
            'referral_code': db_user.referral_code,
            'error': False
        }
        
        return JsonResponse(data)
        
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': True, 'message': 'Пользователь не найден'}, status=404)
    except Exception as e:
        logger.error(f"Ошибка при получении деталей реферала: {e}")
        return JsonResponse({'error': True, 'message': str(e)}, status=500)








######################################################################################

MONTHS_RU = {
    1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
    5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
    9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь',
}


@login_required
def monthly_reports_history(request):
    """История месячных отчётов — только для магазинов и администраторов"""
    from django.core.exceptions import PermissionDenied

    if not (request.user.is_store or request.user.is_staff):
        raise PermissionDenied("У вас нет доступа к истории отчётов")

    year_filter  = request.GET.get('year', '').strip()
    month_filter = request.GET.get('month', '').strip()
    search       = request.GET.get('search', '').strip()

    reports = MonthlyReport.objects.select_related('user').order_by('-year', '-month', 'user__username')

    if year_filter:
        reports = reports.filter(year=year_filter)
    if month_filter:
        reports = reports.filter(month=month_filter)
    if search:
        reports = reports.filter(
            user__username__icontains=search
        ) | reports.filter(
            user__first_name__icontains=search
        ) | reports.filter(
            user__last_name__icontains=search
        )

    available_years = (
        MonthlyReport.objects.values_list('year', flat=True)
        .distinct().order_by('-year')
    )

    paginator = Paginator(reports, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'reports': page_obj,
        'available_years': available_years,
        'months': [(k, v) for k, v in MONTHS_RU.items()],
        'year_filter': year_filter,
        'month_filter': month_filter,
        'search': search,
        'total_count': reports.count(),
    }
    return render(request, 'cabinet/monthly_reports.html', context)


@login_required
def export_monthly_reports_excel(request):
    """Экспорт месячных отчётов в Excel (.xlsx)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.core.exceptions import PermissionDenied

    if not (request.user.is_store or request.user.is_staff):
        raise PermissionDenied("У вас нет доступа")

    year_filter  = request.GET.get('year', '').strip()
    month_filter = request.GET.get('month', '').strip()
    search       = request.GET.get('search', '').strip()

    reports = MonthlyReport.objects.select_related('user').order_by('-year', '-month', 'user__username')
    if year_filter:
        reports = reports.filter(year=year_filter)
    if month_filter:
        reports = reports.filter(month=month_filter)
    if search:
        reports = reports.filter(
            user__username__icontains=search
        ) | reports.filter(
            user__first_name__icontains=search
        ) | reports.filter(
            user__last_name__icontains=search
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Месячные отчёты'

    headers = [
        'Пользователь', 'Год', 'Месяц', 'Уровень партнёра',
        'LO (личный объём)', 'GO (групп. объём)', 'Боковой объём',
        'Баллы', 'Вероны',
        'Личный бонус', 'Структурный бонус', 'Менторский бонус',
        'Доп. бонус', 'Суммарный бонус',
        'Личный доход', 'Групповой доход', 'Лидерский доход',
        'Доход с бок. объёма', 'Итого доход', 'Общий доход',
        'Новые рефералы', 'Активные рефералы',
        'Кол-во заказов', 'Сумма заказов',
        'Дата создания',
    ]

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    for report in reports:
        ws.append([
            report.user.username,
            report.year,
            MONTHS_RU.get(report.month, report.month),
            report.partner_level,
            float(report.personal_volume),
            float(report.group_volume),
            float(report.side_volume),
            float(report.points),
            float(report.veron),
            float(report.personal_bonus),
            float(report.structure_bonus),
            float(report.mentor_bonus),
            report.extra_bonus,
            float(report.bonus_total),
            float(report.personal_money),
            float(report.group_money),
            float(report.leader_money),
            float(report.side_vol_money),
            float(report.total_money),
            float(report.total_income),
            report.new_referrals,
            report.active_referrals_count,
            report.purchases_count,
            float(report.purchases_amount),
            report.created_at.strftime('%d.%m.%Y %H:%M'),
        ])

    # Авто-ширина столбцов
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 32)

    filename_parts = ['monthly_reports']
    if year_filter:
        filename_parts.append(year_filter)
    if month_filter:
        filename_parts.append(month_filter)
    filename = '_'.join(filename_parts) + '.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


######################################################################################

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash

from .forms import ProfileUpdateForm, CustomSetPasswordForm


@login_required
def settings_view(request):
    user = request.user

    if request.method == 'POST':
        profile_form = ProfileUpdateForm(request.POST, instance=user)
        password_form = CustomSetPasswordForm(user, request.POST)

        if 'update_profile' in request.POST:
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, 'Профиль обновлен')
                return redirect('cabinet:settings')

        elif 'change_password' in request.POST:
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Пароль успешно изменён')
                return redirect('cabinet:settings')

    else:
        profile_form = ProfileUpdateForm(instance=user)
        password_form = CustomSetPasswordForm(user)

    return render(request, 'cabinet/settings.html', {
        'profile_form': profile_form,
        'password_form': password_form,
    })




# @login_required
# def generate_monthly_report(request):
#     if request.method != 'POST':
#         return JsonResponse({'error': 'Метод не поддерживается'}, status=405)

#     from datetime import date
#     from django.db.models import Sum, Count
#     from concurrent.futures import ThreadPoolExecutor, as_completed

#     today = date.today()
#     year = today.year
#     month = today.month

#     # 1. Обходим всё дерево структуры
#     all_users = []
#     visited = set()

#     # ── Предзагружаем всех, кто УЖЕ получал extra_bonus ранее ──────
#     # Исключаем текущий месяц — смотрим только прошлые периоды
#     users_with_extra_bonus = set(
#         MonthlyReport.objects.filter(
#             extra_bonus__isnull=False
#         ).exclude(
#             extra_bonus=''
#         ).exclude(
#             year=year, month=month  # текущий месяц не считаем "ранее"
#         ).values_list('user_id', flat=True)
#     )  # ← закрывающая скобка была потеряна в оригинале

#     def traverse(user):
#         if user.pk in visited:
#             return
#         visited.add(user.pk)
#         all_users.append(user)
#         for referral in CustomUser.objects.filter(referrer=user):
#             traverse(referral)

#     traverse(request.user)

#     # 2. Параллельно запрашиваем /status из FastAPI
#     def fetch_status(user):
#         try:
#             resp = requests.get(
#                 f"{FASTAPI_SERVICE_URL}/user/users/{user.username}/status",
#                 timeout=5
#             )
#             if resp.status_code == 200:
#                 data = resp.json()
#                 if not data.get('error'):
#                     return user.pk, data.get('data', {})
#         except Exception as e:
#             logger.warning(f"Не удалось получить статус для {user.username}: {e}")
#         return user.pk, {}

#     status_map = {}
#     with ThreadPoolExecutor(max_workers=10) as executor:
#         futures = {executor.submit(fetch_status, u): u for u in all_users}
#         for future in as_completed(futures):
#             pk, status = future.result()
#             status_map[pk] = status

#     # 3. Сохраняем MonthlyReport
#     saved = 0
#     errors = 0

#     for user in all_users:
#         st = status_map.get(user.pk, {})

#         # ── Если пользователь уже получал extra_bonus — не сохраняем ──
#         already_received = user.pk in users_with_extra_bonus
#         extra_bonus_value = '' if already_received else (st.get('extra_bonus', '') or '')

#         purchases_agg = Purchase.objects.filter(
#             user=user, date__year=year, date__month=month
#         ).aggregate(
#             count=Count('id'),
#             amount=Sum('amount'),
#         )

#         new_referrals = CustomUser.objects.filter(
#             referrer=user,
#             date_joined__year=year,
#             date_joined__month=month,
#         ).count()

#         try:
#             MonthlyReport.objects.update_or_create(
#                 user=user,
#                 year=year,
#                 month=month,
#                 defaults={
#                     'personal_volume':        st.get('lo', 0) or 0,
#                     'group_volume':           st.get('go', 0) or 0,
#                     'side_volume':            st.get('side_volume', 0) or 0,
#                     'points':                 st.get('points', 0) or 0,
#                     'veron':                  st.get('veron', 0) or 0,
#                     'personal_bonus':         st.get('personal_bonus', 0) or 0,
#                     'structure_bonus':        st.get('structure_bonus', 0) or 0,
#                     'mentor_bonus':           st.get('mentor_bonus', 0) or 0,
#                     'extra_bonus':            extra_bonus_value,  # ← проверенное значение
#                     'bonus_total':            sum(float(st.get(k, 0) or 0) for k in ('personal_bonus', 'structure_bonus', 'mentor_bonus')),
#                     'personal_money':         st.get('personal_money', 0) or 0,
#                     'group_money':            st.get('group_money', 0) or 0,
#                     'leader_money':           st.get('leader_money', 0) or 0,
#                     'side_vol_money':         st.get('side_vol_money', 0) or 0,
#                     'total_money':            st.get('total_money', 0) or 0,
#                     'total_income':           st.get('total_income', 0) or 0,
#                     'partner_level':          st.get('qualification', '') or '',
#                     'new_referrals':          new_referrals,
#                     'active_referrals_count': user.active_referrals,
#                     'purchases_count':        purchases_agg['count'] or 0,
#                     'purchases_amount':       purchases_agg['amount'] or 0,
#                 }
#             )
#             saved += 1
#         except Exception as e:
#             logger.error(f"Ошибка сохранения отчёта для {user.username}: {e}")
#             errors += 1

#     # ... остальной код (reset) без изменений ...

#     logger.info(f"MonthlyReport: сохранено {saved}, ошибок {errors} ({month}/{year})")

#     # 4. Параллельно сбрасываем данные в FastAPI для каждого пользователя
#     def reset_user_data(user):
#         try:
#             resp = requests.post(
#                 f"{FASTAPI_SERVICE_URL}/user/users/{user.username}/reset",
#                 timeout=5
#             )
#             if resp.status_code == 200:
#                 payload = resp.json()
#                 if not payload.get('error'):
#                     return user.pk, True
#         except Exception as e:
#             logger.warning(f"Не удалось сбросить данные для {user.username}: {e}")
#         return user.pk, False

#     reset_ok = 0
#     reset_errors = 0

#     with ThreadPoolExecutor(max_workers=10) as executor:
#         reset_futures = {executor.submit(reset_user_data, u): u for u in all_users}
#         for future in as_completed(reset_futures):
#             _pk, success = future.result()
#             if success:
#                 reset_ok += 1
#             else:
#                 reset_errors += 1

#     logger.info(f"FastAPI reset: успешно {reset_ok}, ошибок {reset_errors} ({month}/{year})")

#     return JsonResponse({
#         'success': True,
#         'saved': saved,
#         'errors': errors,
#         'month': month,
#         'year': year,
#         'reset_ok': reset_ok,
#         'reset_errors': reset_errors,
#     })




# @login_required
# def refresh_stats(request):
#     """Обновление статистики по AJAX запросу"""
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             force_refresh = data.get('force', False)
#         except:
#             force_refresh = False
#
#     service = FastAPIService()
#     user_stats = service.get_user_stats(request.user.username, force_refresh=force_refresh)
#
#     if user_stats is None:
#         # Запускаем асинхронное обновление
#         task = update_user_stats_cache.delay(request.user.username)
#         return JsonResponse({
#             'status': 'updating',
#             'message': 'Данные обновляются...',
#             'task_id': str(task.id)
#         })
#
#     return JsonResponse({
#         'status': 'success',
#         'data': user_stats,
#         'refreshed': True
#     })
#
#
# @login_required
# def get_stats_json(request):
#     user_id = request.user.username
#     cache_key = f"user:stats:{user_id}"
#
#     stats = cache.get(cache_key)
#
#     if not stats:
#         update_user_stats_cache.delay(user_id)
#         return JsonResponse(
#             {"status": "loading"},
#             status=202
#         )
#
#     return JsonResponse(
#         {"status": "ok", "data": stats}
#     )